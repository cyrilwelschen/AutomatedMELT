"""
Collection of functions to perform power line analysis
"""

from ctypes import *
from dwfconstants import *
import os
import time
import sys
import math # for imp_analyser
import matplotlib.pyplot as plt
import numpy as np
from datetime import datetime as dt
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.chart import (
    LineChart,
    Reference,
)
from openpyxl.chart.axis import DateAxis

if sys.platform.startswith("win"):
    dwf = cdll.dwf
elif sys.platform.startswith("darwin"):
    dwf = cdll.LoadLibrary("/Library/Frameworks/dwf.framework/dwf")
else:
    dwf = cdll.LoadLibrary("libdwf.so")

def write(ws, headers, *data):
    ws.append(headers)
    for row in zip(*data):
        ws.append(row)

def make_chart(ws_data=None, ws_graph=None, ws_grap_startfield="A1", min_col=2, min_row=1, max_col=4, max_row=12, y_name="Y-Axis", x_name="X-Axis"):
    assert ws_data and ws_graph, "Excel Worksheets missing"

class AnaConnection:
    def __init__(self, plot=False, excel=False):
        self.plot = plot
        self.excel = excel
        self.wb = None
        self.wb_path = None
        self.file_name = None
        self.ws_voltage = None
        self.ws_impedance = None
        self.ws_overview = None
        print("Creating AD2 logical connection instance")

    def create_wb(self):
        self.wb = Workbook()
        ws = self.wb.active
        self.ws_overview = self.wb.create_sheet("Overview", 0)
        self.ws_voltage = self.wb.create_sheet("Voltages", 1)
        self.ws_impedance = self.wb.create_sheet("Impedance", 2)
        self.wb.remove(ws)
        working_directory = """O:\Admin\INI\RLA\ADP\QI\\0_Public\RFT-V Messgerät\AD2_automated_MELT_Results"""
        # test permission
        if not os.path.exists(working_directory):
            working_directory = "./"
        line_name = "line_name"
        self.file_name = '{}_{}.xlsx'.format(dt.now().strftime("%Y%m%d_%H%M%S"), line_name)
        self.wb_path = os.path.join(working_directory, self.file_name) 
        self.ws_overview["A1"] = "Filename: {}".format(self.file_name)
        self.wb.save(self.wb_path)

    def dwf_version(self):
        #print(DWF version)
        version = create_string_buffer(16)
        dwf.FDwfGetVersion(version)
        print("DWF Version: "+str(version.value))

    def connect(self):
        self.excel_filename = self.file_name
        print("Establish AD2 connection")
        self.hdwf = c_int()
        dwf.FDwfDeviceOpen(c_int(-1), byref(self.hdwf))
        if self.hdwf.value == hdwfNone.value:
            szerr = create_string_buffer(512)
            dwf.FDwfGetLastErrorMsg(szerr)
            print(szerr.value)
            print("failed to open device")
            return False
        else:
            print("successfully connected to AD2")
            return True

    def disconnect(self):
        print("Disconnecting from all devices")
        dwf.FDwfDeviceCloseAll()

    def impedance(self, steps=100, start=1e2, stop=1e6, reference=1e2):
        dwf.FDwfDeviceAutoConfigureSet(self.hdwf, c_int(3)) 
        # vars
        sts = c_byte()
        rgHz = [0.0]*steps
        rgRs = [0.0]*steps
        rgXs = [0.0]*steps
        # setup
        print("Reference: "+str(reference)+" Ohm  Frequency: "+str(start)+" Hz ... "+str(stop/1e3)+" kHz for nanofarad capacitors")
        dwf.FDwfAnalogImpedanceReset(self.hdwf)
        dwf.FDwfAnalogImpedanceModeSet(self.hdwf, c_int(8)) # 0 = W1-C1-DUT-C2-R-GND, 1 = W1-C1-R-C2-DUT-GND, 8 = AD IA adapter
        dwf.FDwfAnalogImpedanceReferenceSet(self.hdwf, c_double(reference)) # reference resistor value in Ohms
        dwf.FDwfAnalogImpedanceFrequencySet(self.hdwf, c_double(start)) # frequency in Hertz
        dwf.FDwfAnalogImpedanceAmplitudeSet(self.hdwf, c_double(1)) # 1V amplitude = 2V peak2peak signal
        dwf.FDwfAnalogImpedanceConfigure(self.hdwf, c_int(1)) # start
        time.sleep(2)

        #measurement
        for i in range(100):
            hz = stop * pow(10.0, 1.0*(1.0*i/(steps-1)-1)*math.log10(stop/start)) # exponential frequency steps
            rgHz[i] = hz
            dwf.FDwfAnalogImpedanceFrequencySet(self.hdwf, c_double(hz)) # frequency in Hertz
            time.sleep(0.01) 
            dwf.FDwfAnalogImpedanceStatus(self.hdwf, None) # ignore last capture since we changed the frequency
            while True:
                if dwf.FDwfAnalogImpedanceStatus(self.hdwf, byref(sts)) == 0:
                    dwf.FDwfGetLastErrorMsg(szerr)
                    print(str(szerr.value))
                    quit()
                if sts.value == 2:
                    break
            resistance = c_double()
            reactance = c_double()
            dwf.FDwfAnalogImpedanceStatusMeasure(self.hdwf, DwfAnalogImpedanceResistance, byref(resistance))
            dwf.FDwfAnalogImpedanceStatusMeasure(self.hdwf, DwfAnalogImpedanceReactance, byref(reactance))
            rgRs[i] = abs(resistance.value) # absolute value for logarithmic plot
            rgXs[i] = abs(reactance.value)
        
        #processing
        fig = plt.plot(rgHz, rgRs, rgHz, rgXs)
        ax = plt.gca()
        ax.set_xscale('log')
        ax.set_yscale('log')
        plt.show()
        dwf.FDwfAnalogImpedanceConfigure(self.hdwf, c_int(0)) # stop
        dwf.FDwfDeviceClose(self.hdwf)

    def imp_mag_phase_cap(self, steps=100, start=1e2, stop=1e6, reference=1e2):
        dwf.FDwfDeviceAutoConfigureSet(self.hdwf, c_int(3)) 
        # vars
        sts = c_byte()
        rgHz = [0.0]*steps # freq
        rgZs = [0.0]*steps # imp mag
        rgCp = [0.0]*steps # parallel capacitance
        rgPs = [0.0]*steps # imp phase
        # setup
        # enable positive supply
        dwf.FDwfAnalogIOChannelNodeSet(self.hdwf, c_int(0), c_int(0), c_double(True)) 
        # set voltage to 5 V
        dwf.FDwfAnalogIOChannelNodeSet(self.hdwf, c_int(0), c_int(1), c_double(5.0)) 
        # enable negative supply
        dwf.FDwfAnalogIOChannelNodeSet(self.hdwf, c_int(1), c_int(0), c_double(True)) 
        # set voltage to -5 V
        dwf.FDwfAnalogIOChannelNodeSet(self.hdwf, c_int(1), c_int(1), c_double(-5.0)) 
        dwf.FDwfAnalogIOEnableSet(self.hdwf, c_int(True))
        print("Reference: "+str(reference)+" Ohm  Frequency: "+str(start)+" Hz ... "+str(stop/1e3)+" kHz for nanofarad capacitors")
        dwf.FDwfAnalogImpedanceReset(self.hdwf)
        dwf.FDwfAnalogImpedanceModeSet(self.hdwf, c_int(1)) # 0 = W1-C1-DUT-C2-R-GND (resistor first), 1 = W1-C1-R-C2-DUT-GND (DUT first), 8 = AD IA adapter
        dwf.FDwfAnalogImpedanceReferenceSet(self.hdwf, c_double(reference)) # reference resistor value in Ohms
        dwf.FDwfAnalogImpedanceFrequencySet(self.hdwf, c_double(start)) # frequency in Hertz
        dwf.FDwfAnalogImpedanceAmplitudeSet(self.hdwf, c_double(1)) # 1V amplitude = 2V peak2peak signal
        dwf.FDwfAnalogImpedanceConfigure(self.hdwf, c_int(1)) # start
        time.sleep(2)

        open_res = c_double()
        open_rea = c_double()
        short_res = c_double()
        short_rea = c_double()
        dwf.FDwfAnalogImpedanceCompGet(self.hdwf, byref(open_res), byref(open_rea), byref(short_res), byref(short_rea))
        print("\nCompensation get:\n", open_res.value, open_rea.value, short_res.value, short_rea.value, "\n")

        #measurement
        for i in range(steps):
            hz = stop * pow(10.0, 1.0*(1.0*i/(steps-1)-1)*math.log10(stop/start)) # exponential frequency steps
            rgHz[i] = hz
            dwf.FDwfAnalogImpedanceFrequencySet(self.hdwf, c_double(hz)) # frequency in Hertz
            time.sleep(0.01) 
            dwf.FDwfAnalogImpedanceStatus(self.hdwf, None) # ignore last capture since we changed the frequency
            while True:
                if dwf.FDwfAnalogImpedanceStatus(self.hdwf, byref(sts)) == 0:
                    dwf.FDwfGetLastErrorMsg(szerr)
                    print(str(szerr.value))
                    quit()
                if sts.value == 2:
                    break
            imp_mag = c_double()
            imp_pha = c_double()
            par_cap = c_double()
            dwf.FDwfAnalogImpedanceStatusMeasure(self.hdwf, DwfAnalogImpedanceImpedance, byref(imp_mag))
            dwf.FDwfAnalogImpedanceStatusMeasure(self.hdwf, DwfAnalogImpedanceImpedancePhase, byref(imp_pha))
            dwf.FDwfAnalogImpedanceStatusMeasure(self.hdwf, DwfAnalogImpedanceParallelCapacitance, byref(par_cap))
            rgZs[i] = abs(imp_mag.value) # absolute value for logarithmic plot
            rgPs[i] = abs(imp_pha.value/math.pi*180)
            rgCp[i] = abs(par_cap.value)
        
        #processing
        if self.excel:
            write(self.ws_impedance, ["Freq [Hz]", "|Z| [Ohm]", "Phase [deg]", "C_p [F]"], rgHz, rgZs, rgPs, rgCp)
            self.wb.save(self.wb_path)
            print("Voltages saved to excel")
        if self.plot:
            fig = plt.plot(rgHz, rgZs, label="Imp |Z| [Ohm]")
            # fig = plt.plot(rgHz, rgCp, label="Parallel Cap C_p [Farad]")
            ax = plt.gca()
            ax.set_xscale('log')
            ax.set_yscale('log')
            plt.legend()
            plt.show()
            fig = plt.plot(rgHz, rgPs)
            ax = plt.gca()
            ax.set_xscale('log')
            plt.show()
        dwf.FDwfAnalogImpedanceConfigure(self.hdwf, c_int(0)) # stop
        dwf.FDwfDeviceClose(self.hdwf)

    def get_voltage(self):
        voltage1 = c_double()
        voltage2 = c_double()
        dwf.FDwfAnalogInChannelEnableSet(self.hdwf, c_int(0), c_bool(True)) 
        dwf.FDwfAnalogInChannelOffsetSet(self.hdwf, c_int(0), c_double(0)) 
        dwf.FDwfAnalogInChannelRangeSet(self.hdwf, c_int(0), c_double(5)) 
        dwf.FDwfAnalogInConfigure(self.hdwf, c_bool(False), c_bool(False)) 
        time.sleep(1)
        dwf.FDwfAnalogInStatus(self.hdwf, False, None) 
        dwf.FDwfAnalogInStatusSample(self.hdwf, c_int(0), byref(voltage1))
        dwf.FDwfAnalogInStatusSample(self.hdwf, c_int(1), byref(voltage2))
        print("Channel 1: {} V  Channel 2: {} V".format(voltage1.value, voltage2.value))
        digits = 2
        return [round(voltage1.value, digits), round(voltage2.value, digits), round(voltage1.value - voltage2.value, digits)]

    def sample_voltage(self, frequency=20000000.0, buffer_size=4000, extension_scaling_factor=1):
        sts = c_byte()
        rgdSamplesCh1 = (c_double*buffer_size)()
        rgdSamplesCh2 = (c_double*buffer_size)()
        cBufMax = c_int()
        dwf.FDwfAnalogInBufferSizeInfo(self.hdwf, 0, byref(cBufMax))
        print("Device buffer size: "+str(cBufMax.value)) 
        dwf.FDwfAnalogInFrequencySet(self.hdwf, c_double(frequency))
        dwf.FDwfAnalogInBufferSizeSet(self.hdwf, c_int(buffer_size)) 
        dwf.FDwfAnalogInChannelEnableSet(self.hdwf, c_int(0), c_bool(True))
        dwf.FDwfAnalogInChannelRangeSet(self.hdwf, c_int(0), c_double(5))
        time.sleep(2)
        print("Starting oscilloscope")
        dwf.FDwfAnalogInConfigure(self.hdwf, c_bool(False), c_bool(True))
        while True:
            print("doing")
            time.sleep(0.3)
            dwf.FDwfAnalogInStatus(self.hdwf, c_int(1), byref(sts))
            if sts.value == DwfStateDone.value :
                break
            time.sleep(0.1)
        print("Acquisition done")
        dwf.FDwfAnalogInStatusData(self.hdwf, 0, rgdSamplesCh1, buffer_size) # get channel 1 data
        dwf.FDwfAnalogInStatusData(self.hdwf, 1, rgdSamplesCh2, buffer_size) # get channel 2 data
        rgdSamplesCh1 = np.array(rgdSamplesCh1) * extension_scaling_factor
        rgdSamplesCh2 = np.array(rgdSamplesCh2) * extension_scaling_factor
        dc1 = sum(rgdSamplesCh1)/len(rgdSamplesCh1)
        dc2 = sum(rgdSamplesCh2)/len(rgdSamplesCh2)
        dc21 = dc1 - dc2
        print([round(dc1, 2), round(dc2, 2), round(dc21, 2)])
        if self.excel:
            write(self.ws_voltage, ["t [s]", "Ch1 [V]", "Ch2 [V]"], np.arange(buffer_size)*1/frequency, rgdSamplesCh1, rgdSamplesCh2)
            self.ws_voltage["E1"] = "Avg. Ch1"
            self.ws_overview["B2"] = "Avg. Ch1"
            self.ws_voltage["E2"] = "Avg. Ch2"
            self.ws_overview["B3"] = "Avg. Ch2"
            self.ws_voltage["E3"] = "AC RMS Ch1"
            self.ws_overview["G2"] = "AC RMS Ch1"
            self.ws_voltage["E4"] = "AC RMS Ch2"
            self.ws_overview["G3"] = "AC RMS Ch2"
            avg_ch1 = np.mean(rgdSamplesCh1)
            avg_ch2 = np.mean(rgdSamplesCh2)
            self.ws_voltage["G1"] = str(avg_ch1) + " V"
            self.ws_overview["C2"] = str(avg_ch1) + " V"
            self.ws_voltage["G2"] = str(avg_ch2) + " V"
            self.ws_overview["C3"] = str(avg_ch2) + " V"
            self.ws_voltage["G3"] = str(np.sqrt(np.sum((rgdSamplesCh1-avg_ch1)**2)/len(rgdSamplesCh1))) + " V"
            self.ws_overview["I2"] = str(np.sqrt(np.sum((rgdSamplesCh1-avg_ch1)**2)/len(rgdSamplesCh1))) + " V"
            self.ws_voltage["G4"] = str(np.sqrt(np.sum((rgdSamplesCh2-avg_ch2)**2)/len(rgdSamplesCh2))) + " V"
            self.ws_overview["I3"] = str(np.sqrt(np.sum((rgdSamplesCh2-avg_ch2)**2)/len(rgdSamplesCh2))) + " V"
            self.wb.save(self.wb_path)
            print("Voltages saved to excel")
        if self.plot:
            plt.plot(np.fromiter(rgdSamplesCh2, dtype = np.float), label="Ch 2")
            plt.plot(np.fromiter(rgdSamplesCh1, dtype = np.float), label="Ch 1")
            plt.legend()
            plt.show()
        return [round(dc1, 2), round(dc2, 2), round(dc21, 2)]

    def resistance_capacitance(self, frequnecy=1e3, reference=1e2):
        """
        param reference: resistor value in Ohms
        param frequency: measurement freq in Hertz
        """
        sts = c_byte()
        capacitance = c_double()
        resistance = c_double()
        reactance = c_double()

        print("Reference: "+str(reference)+" Ohm  Frequency: "+str(frequnecy/1e3)+" kHz for nanofarad capacitors")
        dwf.FDwfAnalogImpedanceReset(self.hdwf)
        dwf.FDwfAnalogImpedanceModeSet(self.hdwf, c_int(8)) # 0 = W1-C1-DUT-C2-R-GND, 1 = W1-C1-R-C2-DUT-GND, 8 = AD IA adapter
        dwf.FDwfAnalogImpedanceReferenceSet(self.hdwf, c_double(reference)) 
        dwf.FDwfAnalogImpedanceFrequencySet(self.hdwf, c_double(frequnecy)) 
        dwf.FDwfAnalogImpedanceAmplitudeSet(self.hdwf, c_double(1))
        dwf.FDwfAnalogImpedanceConfigure(self.hdwf, c_int(1)) # start
        time.sleep(1)

        dwf.FDwfAnalogImpedanceStatus(self.hdwf, None) # ignore last capture, force a new one

        for i in range(10):
            while True:
                if dwf.FDwfAnalogImpedanceStatus(self.hdwf, byref(sts)) == 0:
                    dwf.FDwfGetLastErrorMsg(szerr)
                    print(str(szerr.value))
                    quit()
                if sts.value == 2:
                    break
            dwf.FDwfAnalogImpedanceStatusMeasure(self.hdwf, DwfAnalogImpedanceResistance, byref(resistance))
            dwf.FDwfAnalogImpedanceStatusMeasure(self.hdwf, DwfAnalogImpedanceReactance, byref(reactance))
            dwf.FDwfAnalogImpedanceStatusMeasure(self.hdwf, DwfAnalogImpedanceSeriesCapactance, byref(capacitance))
            print(str(i)+" Resistance: "+str(resistance.value)+" Ohm  Reactance: "+str(reactance.value/1e3)+" kOhm  Capacitance: "+str(capacitance.value*1e9)+" nF")
            time.sleep(0.2)

        dwf.FDwfAnalogImpedanceConfigure(self.hdwf, c_int(0)) # stop
        dwf.FDwfDeviceClose(self.hdwf)

    def make_plots(self):
        c1 = LineChart()
        c1.style = 13
        c1.y_axis.title = '|Z| [Ohm]'
        c1.y_axis.scaling.logBase = 10
        c1.x_axis.title = 'Frequency [Hz]'
        c1.width = 37
        c1.height = 9.5
        data = Reference(self.ws_impedance, min_col=2, min_row=1, max_row=101)
        c1.add_data(data, titles_from_data=True)
        # Style the lines
        for i in range(len(c1.series)):
            sx = c1.series[i]
            sx.smooth = True
        x_axis = Reference(self.ws_impedance, min_col=1, min_row=1, max_row=101)
        c1.set_categories(x_axis)
        # second chart
        c2 = LineChart()
        c2.y_axis.scaling.logBase = 10
        c2.y_axis.majorGridlines = None
        data = Reference(self.ws_impedance, min_col=4, min_row=1, max_row=101)
        c2.add_data(data, titles_from_data=True)
        c2.y_axis.axId = 200
        c2.y_axis.title = 'C_p [F]'
        # Style the lines
        for i in range(len(c2.series)):
            sx = c2.series[i]
            sx.smooth = True
        # x_axis = Reference(self.ws_impedance, min_col=1, min_row=1, max_row=101)
        # c2.set_categories(x_axis)
        c2.y_axis.crosses = "max"
        c1 += c2
        self.ws_overview.add_chart(c1, "A5")
        # second plot
        c1 = LineChart()
        c1.style = 13
        c1.y_axis.title = 'Phase [degree]'
        c1.x_axis.title = 'Frequency [Hz]'
        c1.width = 37
        c1.height = 9.5
        data = Reference(self.ws_impedance, min_col=3, min_row=1, max_row=101)
        c1.add_data(data, titles_from_data=True)
        # Style the lines
        for i in range(len(c1.series)):
            sx = c1.series[i]
            sx.smooth = True
        x_axis = Reference(self.ws_impedance, min_col=1, min_row=1, max_row=101)
        c1.set_categories(x_axis)
        self.ws_overview.add_chart(c1, "A24")
        self.wb.save(self.wb_path)
        
    def make_voltage_plots(self, buffer_size):
        v_chart = LineChart()
        v_chart.style = 13
        v_chart.y_axis.title = 'Voltage [V]'
        v_chart.x_axis.title = 'time [s]'
        v_chart.width = 37
        v_chart.height = 9.5
        voltages = Reference(self.ws_voltage, min_col=2, max_col=3, min_row=1, max_row=buffer_size)
        v_chart.add_data(voltages, titles_from_data=True)
        for i in range(len(v_chart.series)):
            sx = v_chart.series[i]
            sx.smooth = True
        x_axis = Reference(self.ws_voltage, min_col=1, min_row=1, max_row=buffer_size)
        v_chart.set_categories(x_axis)
        self.ws_overview.add_chart(v_chart, "A43")
        self.wb.save(self.wb_path)

def ask_continue():
    ret = "N"
    while ret != "y" and ret != "Y":
        ret = input("Switch position... Ready? Type 'y' to continue ")
    return True